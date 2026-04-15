#!/usr/bin/env python3
"""
New Relic to Dynatrace Migration Tool

A comprehensive tool for migrating monitoring configurations from
New Relic to Dynatrace.

Usage:
    python migrate.py --full                    # Full migration
    python migrate.py --export-only             # Export from New Relic only
    python migrate.py --import-only --input ./  # Import to Dynatrace only
    python migrate.py --components dashboards   # Migrate specific components
    python migrate.py --dry-run                 # Validate without applying
"""

import json
import os
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional

import click
import structlog
from dotenv import load_dotenv
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn
from rich.table import Table

# Configure logging
structlog.configure(
    processors=[
        structlog.stdlib.filter_by_level,
        structlog.stdlib.add_logger_name,
        structlog.stdlib.add_log_level,
        structlog.processors.TimeStamper(fmt="iso"),
        structlog.dev.ConsoleRenderer()
    ],
    wrapper_class=structlog.stdlib.BoundLogger,
    context_class=dict,
    logger_factory=structlog.stdlib.LoggerFactory(),
    cache_logger_on_first_use=True,
)

logger = structlog.get_logger()
console = Console()

# Load environment variables
load_dotenv()

# Import project modules
from clients import DynatraceClient, NewRelicClient
from config import (
    AVAILABLE_COMPONENTS,
    COMPONENT_DEPENDENCIES,
    get_settings,
)
from transformers import (
    AlertTransformer,
    DashboardTransformer,
    DropRuleTransformer,
    InfrastructureTransformer,
    LogParsingTransformer,
    SLOTransformer,
    SyntheticTransformer,
    TagTransformer,
    WorkloadTransformer,
)


class MigrationOrchestrator:
    """
    Orchestrates the complete migration process from New Relic to Dynatrace.
    """

    def __init__(
        self,
        newrelic_client: Optional[NewRelicClient] = None,
        dynatrace_client: Optional[DynatraceClient] = None,
        output_dir: str = "./output",
        dry_run: bool = False,
        incremental_state: Optional[Any] = None,
        checkpoint: Optional[Any] = None,
        legacy_mode: bool = False,
        canary_plan: Optional[Any] = None,  # migration.canary.CanaryPlan
    ):
        self.nr_client = newrelic_client
        self.dt_client = dynatrace_client
        self.output_dir = Path(output_dir)
        self.dry_run = dry_run
        self.inc_state = incremental_state
        self.checkpoint = checkpoint
        self.legacy_mode = legacy_mode
        # Phase 20: CanaryPlan controls two-wave import. Default (None) ->
        # no-op split so existing behavior is unchanged.
        from migration.canary import CanaryPlan
        self.canary_plan = canary_plan or CanaryPlan()

        # Initialize transformers. Legacy mode swaps in the Gen2 implementations
        # that emit Alerting Profiles, Management Zones, Auto-Tag Rules, etc.
        if legacy_mode:
            from transformers.legacy import (
                LegacyAlertTransformer,
                LegacyDashboardTransformer,
                LegacyDropRuleTransformer,
                LegacyInfrastructureTransformer,
                LegacyLogParsingTransformer,
                LegacySLOTransformer,
                LegacySyntheticTransformer,
                LegacyTagTransformer,
                LegacyWorkloadTransformer,
            )
            logger.warning(
                "Running in Gen2 compatibility mode (--legacy). "
                "Gen3 features (Workflows, OpenPipeline, Segments, Document API) "
                "will NOT be used. This mode is a stop-gap for classic tenants."
            )
            self.dashboard_transformer = LegacyDashboardTransformer()
            self.alert_transformer = LegacyAlertTransformer()
            self.synthetic_transformer = LegacySyntheticTransformer()
            self.slo_transformer = LegacySLOTransformer()
            self.workload_transformer = LegacyWorkloadTransformer()
            self.infrastructure_transformer = LegacyInfrastructureTransformer()
            self.log_parsing_transformer = LegacyLogParsingTransformer()
            self.tag_transformer = LegacyTagTransformer()
            self.drop_rule_transformer = LegacyDropRuleTransformer()
        else:
            self.dashboard_transformer = DashboardTransformer()
            self.alert_transformer = AlertTransformer()
            self.synthetic_transformer = SyntheticTransformer()
            self.slo_transformer = SLOTransformer()
            self.workload_transformer = WorkloadTransformer()
            self.infrastructure_transformer = InfrastructureTransformer()
            self.log_parsing_transformer = LogParsingTransformer()
            self.tag_transformer = TagTransformer()
            self.drop_rule_transformer = DropRuleTransformer()

        # Create output directories
        self.output_dir.mkdir(parents=True, exist_ok=True)
        (self.output_dir / "exports").mkdir(exist_ok=True)
        (self.output_dir / "transformed").mkdir(exist_ok=True)
        (self.output_dir / "reports").mkdir(exist_ok=True)

    def run_full_migration(self, components: List[str]) -> Dict[str, Any]:
        """Run the complete migration process."""
        console.print("\n[bold blue]Starting New Relic to Dynatrace Migration[/bold blue]\n")

        results = {
            "start_time": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "components": {},
            "summary": {
                "total_exported": 0,
                "total_transformed": 0,
                "total_imported": 0,
                "total_warnings": 0,
                "total_errors": 0
            }
        }

        # Resolve component dependencies
        ordered_components = self._resolve_dependencies(components)
        console.print(f"Components to migrate: {', '.join(ordered_components)}\n")

        # Phase 1: Export from New Relic
        console.print("[bold]Phase 1: Exporting from New Relic[/bold]")
        export_data = self._export_phase(ordered_components)
        results["export_data"] = export_data

        # Phase 2: Transform to Dynatrace format
        console.print("\n[bold]Phase 2: Transforming to Dynatrace format[/bold]")
        transformed_data = self._transform_phase(export_data, ordered_components)
        results["transformed_data"] = transformed_data

        # Phase 3: Import to Dynatrace
        if not self.dry_run:
            console.print("\n[bold]Phase 3: Importing to Dynatrace[/bold]")
            import_results = self._import_phase(transformed_data, ordered_components)
            results["import_results"] = import_results
        else:
            console.print("\n[yellow]Phase 3: Skipped (dry run mode)[/yellow]")

            # Dry-run preview
            self._show_preview(transformed_data)

            # Save preview JSON
            preview_dir = self.output_dir / "preview"
            preview_dir.mkdir(exist_ok=True)
            preview_file = preview_dir / "transformed_preview.json"
            with open(preview_file, "w") as f:
                json.dump({k: v for k, v in transformed_data.items()
                          if k not in ("warnings", "errors")}, f, indent=2, default=str)
            console.print(f"\n[dim]Preview saved to {preview_file}[/dim]")

        # Generate report
        results["end_time"] = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
        self._generate_report(results)

        return results

    def _resolve_dependencies(self, components: List[str]) -> List[str]:
        """Resolve component dependencies and return ordered list."""
        ordered = []
        visited = set()

        def visit(component: str):
            if component in visited:
                return
            visited.add(component)

            # Visit dependencies first
            deps = COMPONENT_DEPENDENCIES.get(component, [])
            for dep in deps:
                if dep in components or dep in AVAILABLE_COMPONENTS:
                    visit(dep)

            ordered.append(component)

        for component in components:
            visit(component)

        return ordered

    def _export_phase(self, components: List[str]) -> Dict[str, Any]:
        """Export data from New Relic."""
        export_data = {}

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            console=console
        ) as progress:

            for component in components:
                task = progress.add_task(f"Exporting {component}...", total=1)

                try:
                    if component == "dashboards":
                        export_data["dashboards"] = self.nr_client.get_all_dashboards()
                    elif component == "alerts":
                        export_data["alert_policies"] = self.nr_client.get_all_alert_policies()
                    elif component == "synthetics":
                        export_data["synthetic_monitors"] = self.nr_client.get_all_synthetic_monitors()
                    elif component == "slos":
                        export_data["slos"] = self.nr_client.get_all_slos()
                    elif component == "workloads":
                        export_data["workloads"] = self.nr_client.get_all_workloads()
                    elif component == "notification_channels":
                        export_data["notification_channels"] = self.nr_client.get_notification_channels()

                    progress.update(task, completed=1)
                    console.print(f"  ✓ Exported {component}")

                except Exception as e:
                    logger.error(f"Failed to export {component}", error=str(e))
                    console.print(f"  ✗ Failed to export {component}: {e}")

        # Save export data
        export_file = self.output_dir / "exports" / "newrelic_export.json"
        with open(export_file, "w") as f:
            json.dump(export_data, f, indent=2, default=str)

        console.print(f"\nExport saved to: {export_file}")
        return export_data

    def _is_entity_changed(self, entity: Dict[str, Any], entity_type: str, index: int) -> bool:
        """Check if an entity has changed since the last incremental run. Returns True if it should be processed."""
        if self.inc_state is None:
            return True
        nr_guid = entity.get("guid", entity.get("id", f"{entity_type}-{index}"))
        if not self.inc_state.has_changed(str(nr_guid), entity):
            return False
        return True

    def _update_entity_hash(self, entity: Dict[str, Any], entity_type: str, index: int) -> None:
        """Update the incremental state hash after processing an entity."""
        if self.inc_state is None:
            return
        nr_guid = entity.get("guid", entity.get("id", f"{entity_type}-{index}"))
        self.inc_state.update(str(nr_guid), entity)

    def _transform_phase(
        self,
        export_data: Dict[str, Any],
        components: List[str]
    ) -> Dict[str, Any]:
        """Transform exported data to Dynatrace format."""
        if self.legacy_mode:
            return self._transform_phase_legacy(export_data, components)
        # Gen3 transformed payload buckets.
        # Each value is the JSON-ready Settings 2.0 envelope, Workflow JSON,
        # or Document API content the corresponding DT client method consumes.
        transformed_data = {
            "dashboards": [],            # Document API content payloads
            "workflows": [],             # Automation API workflow JSON
            "anomaly_detectors": [],     # builtin:davis.anomaly-detectors envelopes
            "synthetic_tests": [],       # builtin:synthetic_test envelopes
            "slos": [],                  # builtin:monitoring.slo envelopes
            "segments": [],              # builtin:segment envelopes
            "iam_policies": [],          # builtin:iam.policy envelopes
            "openpipeline_processors": [],  # builtin:openpipeline.* envelopes
            "warnings": [],
            "errors": [],
            "skipped": [],
        }

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            console=console
        ) as progress:

            # Transform dashboards
            if "dashboards" in components and "dashboards" in export_data:
                task = progress.add_task("Transforming dashboards...", total=1)
                skipped_count = 0
                items_to_transform = []
                for i, item in enumerate(export_data["dashboards"]):
                    if self._is_entity_changed(item, "dashboard", i):
                        items_to_transform.append((i, item))
                    else:
                        skipped_count += 1
                        transformed_data["skipped"].append({"type": "dashboard", "name": item.get("name", ""), "reason": "unchanged"})
                if items_to_transform:
                    results = self.dashboard_transformer.transform_all(
                        [item for _, item in items_to_transform]
                    )
                    for (idx, item), result in zip(items_to_transform, results):
                        if result.success:
                            transformed_data["dashboards"].extend(result.data)
                            self._update_entity_hash(item, "dashboard", idx)
                        transformed_data["warnings"].extend(result.warnings or [])
                        transformed_data["errors"].extend(result.errors or [])
                progress.update(task, completed=1)
                msg = f"  ✓ Transformed {len(transformed_data['dashboards'])} dashboards"
                if skipped_count:
                    msg += f" [dim](skipped {skipped_count} unchanged)[/dim]"
                console.print(msg)

            # Transform alerts
            if "alerts" in components and "alert_policies" in export_data:
                task = progress.add_task("Transforming alerts...", total=1)
                skipped_count = 0
                items_to_transform = []
                for i, item in enumerate(export_data["alert_policies"]):
                    if self._is_entity_changed(item, "alert", i):
                        items_to_transform.append((i, item))
                    else:
                        skipped_count += 1
                        transformed_data["skipped"].append({"type": "alert", "name": item.get("name", ""), "reason": "unchanged"})
                if items_to_transform:
                    results = self.alert_transformer.transform_all(
                        [item for _, item in items_to_transform]
                    )
                    for (idx, item), result in zip(items_to_transform, results):
                        if result.success:
                            if result.workflow:
                                transformed_data["workflows"].append(result.workflow)
                            transformed_data["anomaly_detectors"].extend(
                                result.anomaly_detectors or []
                            )
                            self._update_entity_hash(item, "alert", idx)
                        transformed_data["warnings"].extend(result.warnings or [])
                        transformed_data["errors"].extend(result.errors or [])
                progress.update(task, completed=1)
                msg = (
                    f"  ✓ Transformed {len(transformed_data['workflows'])} workflows, "
                    f"{len(transformed_data['anomaly_detectors'])} Davis anomaly detectors"
                )
                if skipped_count:
                    msg += f" [dim](skipped {skipped_count} unchanged)[/dim]"
                console.print(msg)

            # Transform synthetic monitors
            if "synthetics" in components and "synthetic_monitors" in export_data:
                task = progress.add_task("Transforming synthetic monitors...", total=1)
                skipped_count = 0
                items_to_transform = []
                for i, item in enumerate(export_data["synthetic_monitors"]):
                    if self._is_entity_changed(item, "synthetic", i):
                        items_to_transform.append((i, item))
                    else:
                        skipped_count += 1
                        transformed_data["skipped"].append({"type": "synthetic", "name": item.get("name", ""), "reason": "unchanged"})
                if items_to_transform:
                    results = self.synthetic_transformer.transform_all(
                        [item for _, item in items_to_transform]
                    )
                    for (idx, item), result in zip(items_to_transform, results):
                        if result.success:
                            transformed_data["synthetic_tests"].append(result.monitor)
                            self._update_entity_hash(item, "synthetic", idx)
                        transformed_data["warnings"].extend(result.warnings or [])
                        transformed_data["errors"].extend(result.errors or [])
                progress.update(task, completed=1)
                msg = (
                    f"  ✓ Transformed {len(transformed_data['synthetic_tests'])} "
                    "synthetic tests (builtin:synthetic_test)"
                )
                if skipped_count:
                    msg += f" [dim](skipped {skipped_count} unchanged)[/dim]"
                console.print(msg)

            # Transform SLOs
            if "slos" in components and "slos" in export_data:
                task = progress.add_task("Transforming SLOs...", total=1)
                skipped_count = 0
                items_to_transform = []
                for i, item in enumerate(export_data["slos"]):
                    if self._is_entity_changed(item, "slo", i):
                        items_to_transform.append((i, item))
                    else:
                        skipped_count += 1
                        transformed_data["skipped"].append({"type": "slo", "name": item.get("name", ""), "reason": "unchanged"})
                if items_to_transform:
                    results = self.slo_transformer.transform_all([item for _, item in items_to_transform])
                    for (idx, item), result in zip(items_to_transform, results):
                        if result.success:
                            transformed_data["slos"].append(result.slo)
                            self._update_entity_hash(item, "slo", idx)
                        transformed_data["warnings"].extend(result.warnings or [])
                        transformed_data["errors"].extend(result.errors or [])
                progress.update(task, completed=1)
                msg = f"  ✓ Transformed {len(transformed_data['slos'])} SLOs"
                if skipped_count:
                    msg += f" [dim](skipped {skipped_count} unchanged)[/dim]"
                console.print(msg)

            # Transform workloads
            if "workloads" in components and "workloads" in export_data:
                task = progress.add_task("Transforming workloads...", total=1)
                skipped_count = 0
                items_to_transform = []
                for i, item in enumerate(export_data["workloads"]):
                    if self._is_entity_changed(item, "workload", i):
                        items_to_transform.append((i, item))
                    else:
                        skipped_count += 1
                        transformed_data["skipped"].append({"type": "workload", "name": item.get("name", ""), "reason": "unchanged"})
                if items_to_transform:
                    results = self.workload_transformer.transform_all(
                        [item for _, item in items_to_transform]
                    )
                    for (idx, item), result in zip(items_to_transform, results):
                        if result.success:
                            if result.segment:
                                transformed_data["segments"].append(result.segment)
                            if result.iam_policy:
                                transformed_data["iam_policies"].append(result.iam_policy)
                            self._update_entity_hash(item, "workload", idx)
                        transformed_data["warnings"].extend(result.warnings or [])
                        transformed_data["errors"].extend(result.errors or [])
                progress.update(task, completed=1)
                msg = (
                    f"  ✓ Transformed {len(transformed_data['segments'])} segments + "
                    f"{len(transformed_data['iam_policies'])} IAM policies"
                )
                if skipped_count:
                    msg += f" [dim](skipped {skipped_count} unchanged)[/dim]"
                console.print(msg)

        # Save transformed data
        transform_file = self.output_dir / "transformed" / "dynatrace_config.json"
        with open(transform_file, "w") as f:
            json.dump(transformed_data, f, indent=2, default=str)

        console.print(f"\nTransformed data saved to: {transform_file}")

        if transformed_data["warnings"]:
            console.print(f"\n[yellow]Warnings: {len(transformed_data['warnings'])}[/yellow]")

        return transformed_data

    def _save_checkpoint_if_needed(self, entity_index: int) -> None:
        """Save checkpoint to disk periodically (every 10 entities) for crash resilience."""
        if self.checkpoint and entity_index > 0 and entity_index % 10 == 0:
            self.checkpoint.save(self.output_dir / ".migration-checkpoint.json")

    # ------------------------------------------------------------------
    # Legacy (Gen2) transform + import paths
    # ------------------------------------------------------------------

    def _transform_phase_legacy(
        self, export_data: Dict[str, Any], components: List[str]
    ) -> Dict[str, Any]:
        """Gen2 transform path — populates Alerting Profiles / Metric Events /
        Management Zones / Auto-Tag Rules / Config v1 dashboards/synthetics
        buckets consumed by `_import_phase_legacy`."""
        transformed = {
            "dashboards": [],
            "alerting_profiles": [],
            "metric_events": [],
            "http_monitors": [],
            "browser_monitors": [],
            "slos": [],
            "management_zones": [],
            "warnings": [],
            "errors": [],
            "skipped": [],
        }

        if "dashboards" in components and "dashboards" in export_data:
            for result in self.dashboard_transformer.transform_all(
                export_data["dashboards"]
            ):
                if result.success:
                    transformed["dashboards"].extend(result.data)
                transformed["warnings"].extend(result.warnings or [])
                transformed["errors"].extend(result.errors or [])

        if "alerts" in components and "alert_policies" in export_data:
            for result in self.alert_transformer.transform_all(
                export_data["alert_policies"]
            ):
                if result.success:
                    if result.alerting_profile:
                        transformed["alerting_profiles"].append(result.alerting_profile)
                    transformed["metric_events"].extend(result.metric_events or [])
                transformed["warnings"].extend(result.warnings or [])
                transformed["errors"].extend(result.errors or [])

        if "synthetics" in components and "synthetic_monitors" in export_data:
            for result in self.synthetic_transformer.transform_all(
                export_data["synthetic_monitors"]
            ):
                if result.success:
                    if result.monitor_type == "HTTP":
                        transformed["http_monitors"].append(result.monitor)
                    else:
                        transformed["browser_monitors"].append(result.monitor)
                transformed["warnings"].extend(result.warnings or [])
                transformed["errors"].extend(result.errors or [])

        if "slos" in components and "slos" in export_data:
            for result in self.slo_transformer.transform_all(
                export_data["slos"]
            ):
                if result.success:
                    transformed["slos"].append(result.slo)
                transformed["warnings"].extend(result.warnings or [])
                transformed["errors"].extend(result.errors or [])

        if "workloads" in components and "workloads" in export_data:
            for result in self.workload_transformer.transform_all(
                export_data["workloads"]
            ):
                if result.success:
                    transformed["management_zones"].append(result.management_zone)
                transformed["warnings"].extend(result.warnings or [])
                transformed["errors"].extend(result.errors or [])

        transform_file = self.output_dir / "transformed" / "dynatrace_config_legacy.json"
        with open(transform_file, "w") as f:
            json.dump(transformed, f, indent=2, default=str)
        console.print(
            f"\n[yellow]Legacy transformed data saved to: {transform_file}[/yellow]"
        )
        return transformed

    def _import_phase_legacy(
        self, transformed_data: Dict[str, Any], components: List[str]
    ) -> Dict[str, Any]:
        """Gen2 import path via LegacyDynatraceV1Client."""
        results = {"successful": [], "failed": [], "skipped": []}
        client = self.dt_client  # LegacyDynatraceV1Client in legacy mode

        def _push(items, fn, type_name):
            for item in items:
                try:
                    r = fn(item)
                    bucket = results["successful"] if r.success else results["failed"]
                    bucket.append(
                        {
                            "type": type_name,
                            "name": r.entity_name,
                            "id": r.dynatrace_id,
                            "error": r.error_message,
                        }
                    )
                except Exception as e:  # noqa: BLE001
                    results["failed"].append({"type": type_name, "error": str(e)})

        if "dashboards" in components:
            _push(transformed_data.get("dashboards", []), client.create_dashboard, "dashboard")
        if "alerts" in components:
            _push(
                transformed_data.get("alerting_profiles", []),
                client.create_alerting_profile,
                "alerting_profile",
            )
            _push(
                transformed_data.get("metric_events", []),
                client.create_metric_event,
                "metric_event",
            )
        if "synthetics" in components:
            _push(
                transformed_data.get("http_monitors", []),
                client.create_http_monitor,
                "http_monitor",
            )
            _push(
                transformed_data.get("browser_monitors", []),
                client.create_browser_monitor,
                "browser_monitor",
            )
        if "slos" in components:
            _push(transformed_data.get("slos", []), client.create_slo, "slo")
        if "workloads" in components:
            _push(
                transformed_data.get("management_zones", []),
                client.create_management_zone,
                "management_zone",
            )
        return results

    def _import_phase(
        self,
        transformed_data: Dict[str, Any],
        components: List[str]
    ) -> Dict[str, Any]:
        """Import transformed data to Dynatrace."""
        if self.legacy_mode:
            return self._import_phase_legacy(transformed_data, components)
        import_results = {
            "successful": [],
            "failed": [],
            "skipped": []
        }

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            console=console
        ) as progress:

            # Import dashboards
            if "dashboards" in components:
                dashboards = transformed_data.get("dashboards", [])
                resume_idx = self.checkpoint.get_resume_index("dashboards") if self.checkpoint else 0
                if resume_idx > 0:
                    console.print(f"  [dim]Resuming dashboards from index {resume_idx}[/dim]")
                task = progress.add_task("Importing dashboards...", total=1)
                for i, dashboard in enumerate(dashboards):
                    if i < resume_idx:
                        continue
                    try:
                        result = self.dt_client.create_dashboard(dashboard)
                        if result.success:
                            import_results["successful"].append({
                                "type": "dashboard",
                                "name": result.entity_name,
                                "id": result.dynatrace_id
                            })
                        else:
                            import_results["failed"].append({
                                "type": "dashboard",
                                "name": result.entity_name,
                                "error": result.error_message
                            })
                    except Exception as e:
                        import_results["failed"].append({
                            "type": "dashboard",
                            "error": str(e)
                        })
                    if self.checkpoint:
                        self.checkpoint.mark_complete("dashboards", i)
                        self._save_checkpoint_if_needed(i)
                progress.update(task, completed=1)

            # ---- Gen3 import paths ----------------------------------------
            #
            # Each loop pushes the corresponding transformed payload type to the
            # appropriate Gen3 endpoint via the composed DynatraceClient:
            #   anomaly_detectors   -> Settings 2.0 (builtin:davis.anomaly-detectors)
            #   workflows           -> Automation API
            #   synthetic_tests     -> Settings 2.0 (builtin:synthetic_test)
            #   slos                -> Settings 2.0 (builtin:monitoring.slo)
            #   segments            -> Settings 2.0 (builtin:segment)
            #   iam_policies        -> Settings 2.0 (builtin:iam.policy)
            #   openpipeline_*      -> Settings 2.0 (builtin:openpipeline.*)
            # ---------------------------------------------------------------

            canary_plan = self.canary_plan

            def _import_one(entity, type_name):
                try:
                    result = fn_holder["fn"](entity)
                    if result.success:
                        import_results["successful"].append({
                            "type": type_name,
                            "name": result.entity_name,
                            "id": result.dynatrace_id,
                        })
                    else:
                        import_results["failed"].append({
                            "type": type_name,
                            "name": result.entity_name,
                            "error": result.error_message,
                        })
                except Exception as e:  # noqa: BLE001
                    import_results["failed"].append({
                        "type": type_name,
                        "error": str(e),
                    })

            fn_holder = {"fn": None}

            def _push(items, label, fn, type_name):
                if not items:
                    return
                fn_holder["fn"] = fn
                task = progress.add_task(f"Importing {label}...", total=1)
                # Phase 20: split into canary + rest waves.
                canary, rest = canary_plan.split(items)
                for entity in canary:
                    _import_one(entity, type_name)
                if rest:
                    proceed = canary_plan.approval_gate(label, len(canary), len(items))
                    if not proceed:
                        import_results["skipped"].append({
                            "type": type_name,
                            "reason": f"Canary halt — operator declined wave 2 for {label}",
                            "remaining": len(rest),
                        })
                    else:
                        for entity in rest:
                            _import_one(entity, type_name)
                progress.update(task, completed=1)

            if "alerts" in components:
                _push(
                    transformed_data.get("anomaly_detectors", []),
                    "Davis anomaly detectors",
                    self.dt_client.create_anomaly_detector,
                    "anomaly_detector",
                )
                _push(
                    transformed_data.get("workflows", []),
                    "automation workflows",
                    self.dt_client.create_workflow,
                    "workflow",
                )

            if "synthetics" in components:
                _push(
                    transformed_data.get("synthetic_tests", []),
                    "synthetic tests",
                    self.dt_client.create_synthetic_test,
                    "synthetic_test",
                )

            if "slos" in components:
                _push(
                    transformed_data.get("slos", []),
                    "SLOs",
                    self.dt_client.create_slo,
                    "slo",
                )

            if "workloads" in components:
                _push(
                    transformed_data.get("segments", []),
                    "segments",
                    self.dt_client.create_segment,
                    "segment",
                )
                _push(
                    transformed_data.get("iam_policies", []),
                    "IAM policies",
                    self.dt_client.create_iam_policy,
                    "iam_policy",
                )

            if "openpipeline" in components or "tags" in components or "logs" in components:
                _push(
                    transformed_data.get("openpipeline_processors", []),
                    "OpenPipeline processors",
                    self.dt_client.create_openpipeline_processor,
                    "openpipeline_processor",
                )

        console.print(
            f"\n[green]Successfully imported: {len(import_results['successful'])}[/green]"
        )
        if import_results["failed"]:
            console.print(
                f"[red]Failed imports: {len(import_results['failed'])}[/red]"
            )

        return import_results

    def _show_preview(self, transformed_data: Dict[str, Any]):
        """Show a summary of what would be created in dry-run mode."""
        preview_table = Table(title="Dry-Run Preview — What Would Be Created")
        preview_table.add_column("Entity Type", style="cyan")
        preview_table.add_column("Count", justify="right", style="green")
        preview_table.add_column("Names (first 5)", style="white")

        type_keys = [
            ("Dashboards", "dashboards", "name"),
            ("Workflows", "workflows", "title"),
            ("Davis Anomaly Detectors", "anomaly_detectors", "value.name"),
            ("Synthetic Tests", "synthetic_tests", "value.name"),
            ("SLOs", "slos", "value.name"),
            ("Segments", "segments", "value.name"),
            ("IAM Policies", "iam_policies", "value.name"),
            ("OpenPipeline Processors", "openpipeline_processors", "value.name"),
        ]

        for label, key, name_path in type_keys:
            items = transformed_data.get(key, [])
            if not items:
                continue
            # Extract names
            names = []
            for item in items[:5]:
                if "." in name_path:
                    parts = name_path.split(".")
                    val = item
                    for p in parts:
                        val = val.get(p, {}) if isinstance(val, dict) else ""
                    names.append(str(val) if val else "?")
                else:
                    names.append(item.get(name_path, "?"))
            suffix = f" (+{len(items) - 5} more)" if len(items) > 5 else ""
            preview_table.add_row(label, str(len(items)), ", ".join(names) + suffix)

        console.print(preview_table)

    def _generate_report(self, results: Dict[str, Any]):
        """Generate a migration report."""
        report_file = self.output_dir / "reports" / f"migration_report_{int(time.time())}.json"

        with open(report_file, "w") as f:
            json.dump(results, f, indent=2, default=str)

        console.print(f"\n[bold]Migration report saved to: {report_file}[/bold]")

        # Print summary table
        table = Table(title="Migration Summary")
        table.add_column("Component", style="cyan")
        table.add_column("Exported", justify="right")
        table.add_column("Transformed", justify="right")
        table.add_column("Imported", justify="right", style="green")

        if "export_data" in results:
            export_data = results["export_data"]
            transformed = results.get("transformed_data", {})
            imported = results.get("import_results", {"successful": []})

            components_data = [
                ("Dashboards", len(export_data.get("dashboards", [])),
                 len(transformed.get("dashboards", [])),
                 sum(1 for i in imported.get("successful", []) if i["type"] == "dashboard")),
                ("Alert Policies → Workflows", len(export_data.get("alert_policies", [])),
                 len(transformed.get("workflows", [])),
                 sum(1 for i in imported.get("successful", []) if i["type"] == "workflow")),
                ("Davis Anomaly Detectors", "-",
                 len(transformed.get("anomaly_detectors", [])),
                 sum(1 for i in imported.get("successful", []) if i["type"] == "anomaly_detector")),
                ("Synthetic Tests", len(export_data.get("synthetic_monitors", [])),
                 len(transformed.get("synthetic_tests", [])),
                 sum(1 for i in imported.get("successful", []) if i["type"] == "synthetic_test")),
                ("SLOs", len(export_data.get("slos", [])),
                 len(transformed.get("slos", [])),
                 sum(1 for i in imported.get("successful", []) if i["type"] == "slo")),
                ("Workloads → Segments", len(export_data.get("workloads", [])),
                 len(transformed.get("segments", [])),
                 sum(1 for i in imported.get("successful", []) if i["type"] == "segment")),
            ]

            for name, exported, transformed_count, imported_count in components_data:
                table.add_row(name, str(exported), str(transformed_count), str(imported_count))

        console.print(table)


# =============================================================================
# CLI Commands
# =============================================================================

@click.command()
@click.option("--full", is_flag=True, help="Run full migration")
@click.option("--export-only", is_flag=True, help="Export from New Relic only")
@click.option("--import-only", is_flag=True, help="Import to Dynatrace only")
@click.option("--input", "input_dir", type=click.Path(), help="Input directory for import-only mode")
@click.option("--output", "output_dir", type=click.Path(), default="./output", help="Output directory")
@click.option("--components", type=str, default=None, help="Comma-separated list of components")
@click.option("--dry-run", is_flag=True, help="Validate without applying changes")
@click.option("--list-components", is_flag=True, help="List available components")
@click.option("--rollback", "rollback_file", type=click.Path(exists=True), help="Rollback migration using manifest file")
@click.option("--resume", is_flag=True, help="Resume from last checkpoint")
@click.option("--incremental", is_flag=True, help="Only migrate changed entities")
@click.option("--report", is_flag=True, help="Generate conversion report after migration")
@click.option("--retry", "retry_file", type=click.Path(exists=True), help="Retry failed entities from previous run")
@click.option("--diff", "show_diff", is_flag=True, help="Compare transformed entities against live DT environment")
@click.option("--legacy", is_flag=True, help="Gen2 compatibility mode for classic tenants (emits Alerting Profiles, Management Zones, Auto-Tag Rules, Config v1 dashboards/synthetics). Prefer Gen3 default.")
@click.option("--canary", "canary_pct", type=float, default=None, help="Two-wave import: push CANARY%% of each entity bucket first, then prompt before continuing. Range 1-99.")
@click.option("--canary-auto-proceed", is_flag=True, help="In canary mode, skip the interactive prompt and proceed automatically (CI / scripted use).")
def main(
    full: bool,
    export_only: bool,
    import_only: bool,
    input_dir: Optional[str],
    output_dir: str,
    components: Optional[str],
    dry_run: bool,
    list_components: bool,
    rollback_file: Optional[str],
    resume: bool,
    incremental: bool,
    report: bool,
    retry_file: Optional[str],
    show_diff: bool,
    legacy: bool,
    canary_pct: Optional[float],
    canary_auto_proceed: bool,
):
    """New Relic to Dynatrace Migration Tool."""

    # Handle rollback (Phase 20: now executes against DT for Gen3 entity types)
    if rollback_file:
        from migration.state import RollbackManifest
        manifest = RollbackManifest.load(Path(rollback_file))
        entries = manifest.get_entries()
        if not entries:
            console.print("[yellow]Manifest is empty — nothing to rollback[/yellow]")
            return
        console.print(f"[bold red]Rollback targets ({len(entries)} entities):[/bold red]")
        for entry in entries[:10]:
            console.print(f"  • {entry['entity_type']}: {entry['name']} ({entry['dynatrace_id']})")
        if len(entries) > 10:
            console.print(f"  ... and {len(entries) - 10} more")
        if dry_run:
            console.print("[yellow]--dry-run set; no deletes executed.[/yellow]")
            return
        if not click.confirm("Proceed with rollback?"):
            console.print("[yellow]Rollback cancelled[/yellow]")
            return

        # Need DT credentials to actually delete.
        try:
            settings_obj = get_settings()
        except Exception as e:  # noqa: BLE001
            console.print(f"[red]Configuration error: {e}[/red]"); sys.exit(1)
        if legacy:
            from clients.legacy import LegacyDynatraceV1Client
            dt_for_rollback = LegacyDynatraceV1Client(
                api_token=settings_obj.dynatrace.api_token,
                environment_url=settings_obj.dynatrace.environment_url,
            )
            console.print("[yellow]Rollback in --legacy mode (Gen2 deletes).[/yellow]")
        else:
            dt_for_rollback = DynatraceClient(
                environment_url=settings_obj.dynatrace.environment_url,
                api_token=settings_obj.dynatrace.api_token,
            )

        successes = failures = 0
        for entry in entries:
            r = dt_for_rollback.delete_entity(
                entry["entity_type"], entry["dynatrace_id"]
            ) if hasattr(dt_for_rollback, "delete_entity") else None
            if r is None:
                console.print(
                    f"[yellow]  · {entry['entity_type']} '{entry['name']}': "
                    "no delete handler on legacy client[/yellow]"
                )
                failures += 1
                continue
            if r.success:
                successes += 1
            else:
                failures += 1
                console.print(
                    f"[red]  · failed {entry['entity_type']} '{entry['name']}': "
                    f"{r.error_message}[/red]"
                )
        console.print(
            f"[bold]Rollback complete:[/bold] [green]{successes} ok[/green], "
            f"[red]{failures} failed[/red]"
        )
        return

    # Handle retry of failed entities
    if retry_file:
        from migration.retry import FailedEntities
        failed = FailedEntities.load(Path(retry_file))
        if failed.is_empty():
            console.print("[yellow]No failed entities to retry[/yellow]")
            return
        console.print(f"[bold]Retrying {len(failed.entries)} failed entities...[/bold]")
        for entry in failed.entries:
            console.print(f"  • {entry['entity_type']}: {entry['name']}")
        console.print("[yellow]Retry requires re-running with transformed data and DT credentials[/yellow]")
        return

    if list_components:
        console.print("\n[bold]Available Components:[/bold]")
        for component in AVAILABLE_COMPONENTS:
            deps = COMPONENT_DEPENDENCIES.get(component, [])
            dep_str = f" (depends on: {', '.join(deps)})" if deps else ""
            console.print(f"  • {component}{dep_str}")
        return

    # Parse components
    if components:
        component_list = [c.strip() for c in components.split(",")]
    else:
        component_list = ["dashboards", "alerts", "synthetics", "slos", "workloads"]

    # Validate environment variables
    try:
        settings = get_settings()
    except Exception as e:
        console.print(f"[red]Configuration error: {e}[/red]")
        console.print("\nPlease set the required environment variables:")
        console.print("  NEW_RELIC_API_KEY")
        console.print("  NEW_RELIC_ACCOUNT_ID")
        console.print("  DYNATRACE_API_TOKEN")
        console.print("  DYNATRACE_ENVIRONMENT_URL")
        sys.exit(1)

    # Initialize clients
    nr_client = None
    dt_client = None

    if not import_only:
        nr_client = NewRelicClient(
            api_key=settings.newrelic.api_key,
            account_id=settings.newrelic.account_id,
            region=settings.newrelic.region
        )

    # CLI flag takes precedence over MIGRATION_LEGACY_MODE env var.
    legacy_mode = legacy or settings.migration.legacy_mode

    if not export_only:
        if legacy_mode:
            from clients.legacy import LegacyDynatraceV1Client
            dt_client = LegacyDynatraceV1Client(
                api_token=settings.dynatrace.api_token,
                environment_url=settings.dynatrace.environment_url,
            )
        else:
            dt_client = DynatraceClient(
                environment_url=settings.dynatrace.environment_url,
                api_token=settings.dynatrace.api_token,
            )

        # Validate Dynatrace connection
        if not dt_client.validate_connection():
            console.print("[red]Failed to connect to Dynatrace. Check your API token and URL.[/red]")
            sys.exit(1)

    # Initialize migration state
    from migration.report import ConversionReport
    from migration.state import IncrementalState, MigrationCheckpoint, RollbackManifest

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    checkpoint = None
    if resume:
        checkpoint_file = output_path / ".migration-checkpoint.json"
        if checkpoint_file.exists():
            checkpoint = MigrationCheckpoint.load(checkpoint_file)
            console.print(f"[green]Resuming from checkpoint: {checkpoint_file}[/green]")
        else:
            console.print("[yellow]No checkpoint found — starting fresh[/yellow]")

    inc_state = None
    if incremental:
        state_file = output_path / ".migration-state.json"
        if state_file.exists():
            inc_state = IncrementalState.load(state_file)
            console.print("[green]Incremental mode: will skip unchanged entities[/green]")
        else:
            inc_state = IncrementalState()

    # Phase 20: assemble canary plan from CLI flags.
    from migration.canary import CanaryPlan, auto_approve_gate, cli_prompt_gate
    canary_plan = CanaryPlan(
        canary_percent=canary_pct,
        approval_gate=auto_approve_gate if canary_auto_proceed else cli_prompt_gate,
    )

    # Create orchestrator
    orchestrator = MigrationOrchestrator(
        newrelic_client=nr_client,
        dynatrace_client=dt_client,
        output_dir=output_dir,
        dry_run=dry_run,
        incremental_state=inc_state,
        checkpoint=checkpoint,
        legacy_mode=legacy_mode,
        canary_plan=canary_plan,
    )

    # Run migration
    if full or (not export_only and not import_only):
        results = orchestrator.run_full_migration(component_list)

        # Save failed entities for retry
        from migration.retry import FailedEntities
        failed = FailedEntities()
        for entry in results.get("import_results", {}).get("failed", []):
            failed.add(entry.get("type", ""), entry.get("name", ""), entry.get("error", ""))
        if not failed.is_empty():
            failed_file = output_path / "failed-entities.json"
            failed.save(failed_file)
            console.print(f"[yellow]{len(failed.entries)} failed entities saved to {failed_file}[/yellow]")
            console.print("[dim]Retry with: python migrate.py migrate --retry failed-entities.json[/dim]")

        # Diff against live environment
        if show_diff:
            registry = _create_registry()
            if registry:
                from migration.diff import DiffReport
                diff = DiffReport.generate_diff(results.get("transformed_data", {}), registry)
                diff_summary = diff.summary()
                diff_table = Table(title="Diff: Transformed vs Live Environment")
                diff_table.add_column("Entity", style="white")
                diff_table.add_column("Action", style="cyan")
                diff_table.add_column("Reason", style="dim")
                for entry in diff.entries:
                    action_style = {"CREATE": "green", "UPDATE": "yellow", "CONFLICT": "red", "ORPHAN": "magenta"}.get(entry.action, "white")
                    diff_table.add_row(f"{entry.entity_type}: {entry.name}", f"[{action_style}]{entry.action}[/{action_style}]", entry.reason)
                console.print(diff_table)
                orphan_count = diff_summary.get('orphans', 0)
                summary_parts = [
                    f"{diff_summary['creates']} create",
                    f"{diff_summary['updates']} update",
                    f"{diff_summary['conflicts']} conflict",
                ]
                if orphan_count:
                    summary_parts.append(f"{orphan_count} orphan")
                console.print(f"\n[bold]Diff: {', '.join(summary_parts)}[/bold]")
            else:
                console.print("[yellow]--diff requires DT credentials for live environment comparison[/yellow]")

        # Save rollback manifest
        manifest_file = output_path / "rollback-manifest.json"
        manifest = RollbackManifest()
        for entry in results.get("import_results", {}).get("successful", []):
            manifest.add(entry["type"], entry.get("id", ""), entry.get("name", ""))
        manifest.save(manifest_file)
        if manifest.get_entries():
            console.print(f"[dim]Rollback manifest saved to {manifest_file}[/dim]")

        # Save checkpoint
        if resume or True:  # Always save checkpoint for future resume
            cp = MigrationCheckpoint()
            for comp in component_list:
                cp.mark_complete(comp, -1)  # Mark all as complete
            cp.save(output_path / ".migration-checkpoint.json")

        # Save incremental state
        if inc_state is not None:
            inc_state.save(output_path / ".migration-state.json")

        # Generate conversion report
        if report:
            report_obj = ConversionReport()
            # Collect query data from transform warnings
            for w in results.get("transformed_data", {}).get("warnings", []):
                if "NRQL" in str(w) or "confidence" in str(w).lower():
                    report_obj.add_query(original_nrql=str(w), converted_dql="", confidence="MEDIUM", warnings=[str(w)])
            report_json = output_path / "reports" / "conversion-report.json"
            report_html = output_path / "reports" / "conversion-report.html"
            report_json.parent.mkdir(parents=True, exist_ok=True)
            report_obj.generate_json(report_json)
            report_obj.generate_html(report_html)
            summary = report_obj.summary()
            console.print(f"\n[bold]Conversion Report:[/bold] {summary['total']} queries — "
                          f"[green]{summary['high_confidence']} high[/green], "
                          f"[yellow]{summary['medium_confidence']} medium[/yellow], "
                          f"[red]{summary['low_confidence']} low[/red]")
            console.print(f"  JSON: {report_json}")
            console.print(f"  HTML: {report_html}")
    elif export_only:
        orchestrator._export_phase(component_list)
    elif import_only:
        if not input_dir:
            console.print("[red]--input is required for import-only mode[/red]")
            sys.exit(1)

        # Load transformed data
        input_path = Path(input_dir) / "transformed" / "dynatrace_config.json"
        if not input_path.exists():
            input_path = Path(input_dir) / "dynatrace_config.json"

        if not input_path.exists():
            console.print(f"[red]Could not find transformed data at {input_path}[/red]")
            sys.exit(1)

        with open(input_path) as f:
            transformed_data = json.load(f)

        orchestrator._import_phase(transformed_data, component_list)


def _display_compile_result(result, show_original: str = None):
    """Display a compile result with Rich formatting."""
    from rich.panel import Panel

    if show_original:
        console.print(f"\n[bold cyan]NRQL:[/bold cyan] {show_original}")

    if result.success:
        console.print(Panel(result.dql, border_style="green", title="DQL", title_align="left"))
        if result.fixes:
            for f in result.fixes:
                console.print(f"  [green]Fix:[/green] {f}")
        if result.warnings:
            for w in result.warnings:
                console.print(f"  [yellow]Warning:[/yellow] {w}")
        console.print(f"  [dim]Confidence: {result.confidence}[/dim]")
    else:
        console.print(f"[red]Error:[/red] {result.error}")


@click.command("compile")
@click.argument("nrql", required=False)
@click.option("--interactive", "-i", is_flag=True, help="Interactive REPL mode")
@click.option("--file", "-f", "input_file", type=click.Path(exists=True), help="Read queries from file")
@click.option("--output", "-o", "output_file", type=click.Path(), help="Write results to file")
@click.option("--validate", "-v", is_flag=True, help="Validate compiled DQL against live DT environment")
def compile_nrql(nrql: Optional[str], interactive: bool, input_file: Optional[str], output_file: Optional[str], validate: bool):
    """Compile NRQL queries to DQL.

    Modes:

      python migrate.py compile "SELECT count(*) FROM Transaction"

      python migrate.py compile --interactive

      python migrate.py compile --file queries.nrql --output results.dql

      python migrate.py compile --validate "SELECT count(*) FROM Transaction"
    """
    from compiler import NRQLCompiler

    if not nrql and not interactive and not input_file:
        ctx = click.get_current_context()
        click.echo(ctx.get_help())
        sys.exit(1)

    # Optionally create registry for live validation
    registry = None
    if validate:
        registry = _create_registry()
        if not registry:
            console.print("[yellow]Warning: Could not create registry for live validation. "
                          "Set DYNATRACE_ENVIRONMENT_URL and DYNATRACE_API_TOKEN in .env[/yellow]")

    compiler = NRQLCompiler()

    # Interactive REPL mode
    if interactive:
        console.print("[bold]NRQL to DQL Compiler — Interactive Mode[/bold]")
        console.print("Enter NRQL queries (type 'quit' to exit, 'ref' for reference)\n")

        while True:
            try:
                nrql_input = console.input("[cyan]NRQL>[/cyan] ")

                if nrql_input.lower() in ("quit", "exit", "q"):
                    break

                if nrql_input.lower() in ("ref", "reference", "help"):
                    _print_reference_table()
                    continue

                if not nrql_input.strip():
                    continue

                result = compiler.compile(nrql_input)
                _display_compile_result(result)
                console.print()

            except (KeyboardInterrupt, EOFError):
                console.print("\n[yellow]Exiting...[/yellow]")
                break

        return

    # Batch file mode
    if input_file:
        with open(input_file, "r") as f:
            queries = [line.strip() for line in f if line.strip() and not line.startswith("#")]

        results = []
        for q in queries:
            result = compiler.compile(q)
            results.append((q, result))
            _display_compile_result(result, show_original=q)
            console.print("─" * 60)

        if output_file:
            with open(output_file, "w") as f:
                for original, result in results:
                    f.write(f"-- Original: {original}\n")
                    if result.success:
                        f.write(f"{result.dql}\n\n")
                    else:
                        f.write(f"-- Error: {result.error}\n\n")
            console.print(f"\n[green]Results saved to {output_file}[/green]")

        console.print(f"\n[bold]Compiled {len(queries)} queries: "
                      f"[green]{sum(1 for _, r in results if r.success)} succeeded[/green], "
                      f"[red]{sum(1 for _, r in results if not r.success)} failed[/red][/bold]")
        return

    # Single query mode
    result = compiler.compile(nrql)

    if result.success:
        console.print(result.dql)
        if result.warnings:
            for w in result.warnings:
                console.print(f"[yellow]Warning:[/yellow] {w}")

        # Live validation against DT environment
        if validate and registry:
            is_valid, error_msg, _ = registry.validate_dql_syntax(result.dql)
            if is_valid is True:
                console.print("[green]Live validation: DQL is valid[/green]")
            elif is_valid is False:
                console.print(f"[red]Live validation failed:[/red] {error_msg}")
            else:
                console.print(f"[yellow]Live validation skipped:[/yellow] {error_msg}")
    else:
        console.print(f"[red]Error:[/red] {result.error}")
        sys.exit(1)


def _create_registry():
    """Create a DTEnvironmentRegistry from environment variables if available."""
    dt_url = os.environ.get("DYNATRACE_ENVIRONMENT_URL", "")
    api_token = os.environ.get("DYNATRACE_API_TOKEN", "")
    oauth_token = os.environ.get("DYNATRACE_OAUTH_TOKEN", "")

    if not dt_url or (not api_token and not oauth_token):
        return None

    try:
        from registry.environment import DTEnvironmentRegistry
        return DTEnvironmentRegistry(dt_url, oauth_token=oauth_token, api_token=api_token)
    except Exception as e:
        logger.warning("Could not create registry", error=str(e))
        return None


@click.command("audit-slos")
def audit_slos():
    """Audit SLOs in the Dynatrace environment for metric validity.

    Checks all SLOs for missing metrics, invalid aggregations, and NRQL syntax
    that wasn't properly converted. Requires DYNATRACE_ENVIRONMENT_URL and
    DYNATRACE_OAUTH_TOKEN environment variables.
    """
    oauth_token = os.environ.get("DYNATRACE_OAUTH_TOKEN", "")
    api_token = os.environ.get("DYNATRACE_API_TOKEN", "")
    dt_url = os.environ.get("DYNATRACE_ENVIRONMENT_URL", "")

    if not dt_url or not oauth_token:
        console.print("[red]Error: DYNATRACE_ENVIRONMENT_URL and DYNATRACE_OAUTH_TOKEN required[/red]")
        console.print("SLO audit requires OAuth token for Platform SLO API access.")
        sys.exit(1)

    from registry.environment import DTEnvironmentRegistry
    from registry.slo_auditor import SLOAuditor

    registry = DTEnvironmentRegistry(dt_url, oauth_token=oauth_token, api_token=api_token)
    auditor = SLOAuditor(dt_url, oauth_token, api_token=api_token, registry=registry)

    console.print("[bold]Auditing SLOs...[/bold]\n")
    results = auditor.audit()

    # Display results
    table = Table(title="SLO Audit Results", show_header=True, header_style="bold cyan")
    table.add_column("SLO", style="white")
    table.add_column("Status", justify="center")
    table.add_column("Issues", style="yellow")

    for slo_result in results:
        status = "[green]OK[/green]" if slo_result.get("valid") else "[red]FAIL[/red]"
        issues = "; ".join(slo_result.get("issues", [])) or "—"
        table.add_row(slo_result.get("name", "Unknown"), status, issues[:80])

    console.print(table)

    valid = sum(1 for r in results if r.get("valid"))
    console.print(f"\n[bold]{valid}/{len(results)} SLOs valid[/bold]")


@click.command("convert")
@click.argument("nrql")
def convert_nrql(nrql: str):
    """Convert NRQL to DQL with full post-processing and auto-fixes.

    Example: python migrate.py convert "SELECT average(duration) FROM Transaction WHERE appName = 'my-api'"
    """
    from transformers.nrql_converter import NRQLtoDQLConverter

    converter = NRQLtoDQLConverter()
    result = converter.convert(nrql, "CLI query")

    console.print(result.dql)
    if result.fixes:
        for f in result.fixes:
            console.print(f"[green]Fix:[/green] {f}")
    if result.warnings:
        for w in result.warnings:
            console.print(f"[yellow]Warning:[/yellow] {w}")
    console.print(f"\n[dim]Confidence: {result.confidence}[/dim]")


def _print_reference_table():
    """Print the NRQL to DQL quick reference table."""
    table = Table(title="NRQL to DQL Quick Reference", show_header=True, header_style="bold cyan")
    table.add_column("NRQL", style="green")
    table.add_column("DQL", style="blue")

    references = [
        ("SELECT * FROM Log", "fetch logs"),
        ("SELECT count(*) FROM Transaction", "fetch ... | summarize count()"),
        ("WHERE field = 'value'", '| filter field == "value"'),
        ("WHERE field LIKE '%pattern%'", "| filter contains(field, \"pattern\")"),
        ("WHERE field IN ('a', 'b')", '| filter in(field, {"a", "b"})'),
        ("WHERE field IS NULL", "| filter isNull(field)"),
        ("FACET fieldName", "| summarize ..., by:{fieldName}"),
        ("SINCE 1 hour ago", "from:now()-1h"),
        ("LIMIT 100", "| limit 100"),
        ("TIMESERIES", "| makeTimeseries ..."),
        ("─" * 35, "─" * 35),
        ("count(*)", "count()"),
        ("average(field)", "avg(field)"),
        ("sum(field)", "sum(field)"),
        ("max(field)", "max(field)"),
        ("min(field)", "min(field)"),
        ("uniqueCount(field)", "countDistinct(field)"),
        ("percentile(field, 95)", "percentile(field, 95)"),
        ("latest(field)", "takeLast(field)"),
        ("earliest(field)", "takeFirst(field)"),
    ]

    for nrql, dql in references:
        table.add_row(nrql, dql)

    console.print(table)


@click.command("reference")
@click.option("--mappings", "-m", is_flag=True, help="Show full mapping tables (aggregations, event types, attributes)")
def reference(mappings: bool):
    """Show NRQL to DQL reference table.

    Use --mappings to display the full mapping dictionaries.
    """
    _print_reference_table()

    if mappings:
        from transformers.nrql_mapping_rules import AGG_MAP, ATTR_MAP, EVENT_TYPE_MAP

        console.print()
        agg_table = Table(title="Aggregation Mappings", show_header=True, header_style="bold cyan")
        agg_table.add_column("NRQL Function", style="green")
        agg_table.add_column("DQL Function", style="blue")
        for nrql_fn, dql_fn in sorted(AGG_MAP.items()):
            agg_table.add_row(nrql_fn, dql_fn)
        console.print(agg_table)

        console.print()
        event_table = Table(title="Event Type Mappings", show_header=True, header_style="bold cyan")
        event_table.add_column("NR Event Type", style="green")
        event_table.add_column("DT Data Object", style="blue")
        for nr_type, dt_type in sorted(EVENT_TYPE_MAP.items()):
            event_table.add_row(nr_type, dt_type)
        console.print(event_table)

        console.print()
        attr_table = Table(title="Attribute Mappings", show_header=True, header_style="bold cyan")
        attr_table.add_column("NR Attribute", style="green")
        attr_table.add_column("DT Attribute", style="blue")
        for nr_attr, dt_attr in sorted(ATTR_MAP.items()):
            attr_table.add_row(nr_attr, dt_attr)
        console.print(attr_table)


@click.command("batch")
@click.option("--file", "-f", "input_file", required=True, type=click.Path(exists=True), help="CSV or Excel file with NRQL column")
@click.option("--output", "-o", "output_file", type=click.Path(), help="Output file (CSV)")
@click.option("--nrql-column", default="nrql", help="Column name containing NRQL queries (default: nrql)")
def batch_compile(input_file: str, output_file: Optional[str], nrql_column: str):
    """Batch compile NRQL queries from a CSV or Excel file.

    Reads a file with an NRQL column, compiles each query to DQL, and writes
    results with DQL, confidence, and warnings columns.

    Example: python migrate.py batch --file queries.csv --output results.csv
    """
    import csv

    from compiler import NRQLCompiler

    compiler = NRQLCompiler()
    input_path = Path(input_file)
    rows = []

    # Read input
    if input_path.suffix in ('.xlsx', '.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(input_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            nrql_idx = headers.index(nrql_column) if nrql_column in headers else 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                nrql = row[nrql_idx] if row[nrql_idx] else ""
                rows.append({"nrql": str(nrql), "original_row": list(row)})
            wb.close()
        except ImportError:
            console.print("[red]Error: openpyxl required for Excel files. Install with: pip install openpyxl[/red]")
            sys.exit(1)
    else:
        # CSV
        with open(input_path, 'r', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                nrql = row.get(nrql_column, "")
                rows.append({"nrql": nrql, "original_row": row})

    if not rows:
        console.print("[yellow]No queries found in input file[/yellow]")
        return

    # Compile each query
    results = []
    for item in rows:
        nrql = item["nrql"].strip()
        if not nrql:
            results.append({"nrql": "", "dql": "", "confidence": "", "warnings": ""})
            continue
        result = compiler.compile(nrql)
        results.append({
            "nrql": nrql,
            "dql": result.dql if result.success else f"ERROR: {result.error}",
            "confidence": result.confidence if result.success else "FAILED",
            "warnings": "; ".join(result.warnings) if result.warnings else "",
        })

    # Display summary
    succeeded = sum(1 for r in results if r["confidence"] not in ("", "FAILED"))
    failed = sum(1 for r in results if r["confidence"] == "FAILED")
    console.print(f"\n[bold]Batch compiled {len(results)} queries: "
                  f"[green]{succeeded} succeeded[/green], "
                  f"[red]{failed} failed[/red][/bold]")

    # Write output
    out_path = output_file or str(input_path.with_suffix('.results.csv'))
    with open(out_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=["nrql", "dql", "confidence", "warnings"])
        writer.writeheader()
        writer.writerows(results)
    console.print(f"[green]Results saved to {out_path}[/green]")


def _get_version():
    try:
        from _version import __version__
        return __version__
    except ImportError:
        return "unknown"


# Create a click group to support both migration and compile subcommands
@click.group(invoke_without_command=True)
@click.version_option(version=_get_version(), prog_name="nr-to-dt-migration")
@click.pass_context
def cli(ctx):
    """New Relic to Dynatrace Migration Tool."""
    if ctx.invoked_subcommand is None:
        click.echo(ctx.get_help())


def _load_transformed(input_dir: str, legacy: bool) -> Dict[str, Any]:
    input_path = Path(input_dir)
    candidates = [
        input_path / "transformed" / (
            "dynatrace_config_legacy.json" if legacy else "dynatrace_config.json"
        ),
        input_path / "transformed" / "dynatrace_config.json",
        input_path / "preview" / "transformed_preview.json",
        input_path / "dynatrace_config.json",
    ]
    for candidate in candidates:
        if candidate.exists():
            with open(candidate) as f:
                return json.load(f)
    console.print(f"[red]Could not find transformed data in {input_dir}[/red]")
    sys.exit(1)


@click.command("export-monaco")
@click.option("--input", "input_dir", required=True, type=click.Path(exists=True), help="Directory with transformed data")
@click.option("--output", "output_dir", type=click.Path(), default="./monaco-output", help="Monaco output directory")
@click.option("--legacy", is_flag=True, help="Emit Gen2 (Config v1) Monaco project instead of Gen3 default.")
def export_monaco(input_dir: str, output_dir: str, legacy: bool):
    """Export transformed data as Monaco config-as-code project."""
    if legacy:
        from exporters.legacy import LegacyMonacoExporter as MonacoExporter
        logger.warning("Monaco export in Gen2 compatibility mode (--legacy).")
    else:
        from exporters.monaco import MonacoExporter

    transformed_data = _load_transformed(input_dir, legacy=legacy)
    summary = MonacoExporter().export(transformed_data, Path(output_dir))

    console.print(f"\n[green]Monaco project exported to {output_dir}[/green]")
    for entity_type, count in summary.items():
        console.print(f"  {entity_type}: {count}")


@click.command("export-terraform")
@click.option("--input", "input_dir", required=True, type=click.Path(exists=True), help="Directory with transformed data")
@click.option("--output", "output_dir", type=click.Path(), default="./terraform-output", help="Terraform output directory")
@click.option("--legacy", is_flag=True, help="Emit Gen2 Terraform resources instead of Gen3 default.")
def export_terraform(input_dir: str, output_dir: str, legacy: bool):
    """Export transformed data as Terraform HCL configuration."""
    if legacy:
        from exporters.legacy import LegacyTerraformExporter as TerraformExporter
        logger.warning("Terraform export in Gen2 compatibility mode (--legacy).")
    else:
        from exporters.terraform import TerraformExporter

    transformed_data = _load_transformed(input_dir, legacy=legacy)
    summary = TerraformExporter().export(transformed_data, Path(output_dir))

    console.print(f"\n[green]Terraform config exported to {output_dir}[/green]")
    for entity_type, count in summary.items():
        console.print(f"  {entity_type}: {count}")


@click.command("preflight")
def preflight():
    """Probe the target Dynatrace tenant for Gen3 API availability.

    Reports Settings 2.0, Document API, and Automation API reachability.
    Suggests `--legacy` if any Gen3 surface is missing.
    """
    try:
        settings = get_settings()
    except Exception as e:  # noqa: BLE001
        console.print(f"[red]Configuration error: {e}[/red]")
        sys.exit(1)

    client = DynatraceClient(
        environment_url=settings.dynatrace.environment_url,
        api_token=settings.dynatrace.api_token,
    )
    report = client.preflight_gen3()

    table = Table(title="Dynatrace Gen3 Preflight")
    table.add_column("API", style="cyan")
    table.add_column("Reachable", justify="center")
    for name in ("settings_v2", "document_api", "automation_api"):
        ok = report.get(name, False)
        table.add_row(name, "[green]yes[/green]" if ok else "[red]no[/red]")
    console.print(table)

    if not all(report.values()):
        console.print(
            "\n[yellow]One or more Gen3 APIs are not reachable. "
            "Consider running with `--legacy` until the tenant is upgraded.[/yellow]"
        )
        sys.exit(1)
    console.print("\n[green]Gen3 APIs reachable — default mode is safe to use.[/green]")


# Register subcommands
cli.add_command(main, "migrate")
cli.add_command(compile_nrql, "compile")
cli.add_command(convert_nrql, "convert")
cli.add_command(reference, "reference")
cli.add_command(audit_slos, "audit-slos")
cli.add_command(batch_compile, "batch")
cli.add_command(export_monaco, "export-monaco")
cli.add_command(export_terraform, "export-terraform")
cli.add_command(preflight, "preflight")


@click.command("agents")
@click.option(
    "--language",
    type=click.Choice(
        ["java", "dotnet", "nodejs", "python", "ruby", "php", "go"],
        case_sensitive=False,
    ),
    required=True,
    help="APM agent language to migrate on the target host(s).",
)
@click.option("--host", "hostname", default="localhost", help="Target hostname.")
@click.option(
    "--phase",
    type=click.Choice(["uninstall", "install-oneagent", "install-otel", "verify", "all"]),
    default="all",
)
@click.option("--dry-run", is_flag=True, help="Print the action plan without executing.")
def agents_cmd(language: str, hostname: str, phase: str, dry_run: bool):
    """Per-language APM agent migration orchestrator.

    Emits an ordered action plan (commands + rollback hooks) for an
    operator or automation layer to execute. Does NOT run shell commands
    itself — the plan is always printed first.
    """
    from agents import SUPPORTED_LANGUAGES
    Agent = SUPPORTED_LANGUAGES[language.lower()]
    orch = Agent()
    host = {"name": hostname, "hostname": hostname}
    phases = (
        ["uninstall", "install-oneagent", "verify"]
        if phase == "all"
        else [phase]
    )
    for p in phases:
        if p == "uninstall":
            r = orch.uninstall_nr(host, dry_run=dry_run)
        elif p == "install-oneagent":
            r = orch.install_oneagent(host, dry_run=dry_run)
        elif p == "install-otel":
            r = orch.install_otel_fallback(host, dry_run=dry_run)
        else:
            r = orch.verify(host)
        if not r.success:
            for w in r.warnings:
                console.print(f"[yellow]{w}[/yellow]")
            for e in r.errors:
                console.print(f"[red]{e}[/red]")
            continue
        table = Table(title=f"{language} / {p} on {hostname}")
        table.add_column("#", style="dim")
        table.add_column("Action ID", style="cyan")
        table.add_column("Description")
        for i, action in enumerate(r.plan.actions, 1):
            table.add_row(str(i), action.id, action.description)
        console.print(table)


@click.command("scan-instrumentation")
@click.option(
    "--file", "input_file", required=True, type=click.Path(exists=True),
    help="Source file to scan for newrelic.*() SDK calls.",
)
@click.option("--output", "output_file", type=click.Path(), help="Write diff to file.")
def scan_instrumentation_cmd(input_file: str, output_file: Optional[str]):
    """Scan a source file for NR SDK calls and print suggested DT replacements."""
    from transformers.custom_instrumentation_translator import (
        CustomInstrumentationTranslator,
    )
    text = Path(input_file).read_text()
    translator = CustomInstrumentationTranslator()
    result = translator.scan_text(text, input_file)
    diff = translator.render_diff(result)
    if output_file:
        Path(output_file).write_text(diff)
        console.print(f"[green]Diff written to {output_file}[/green]")
    else:
        console.print(diff or "[dim]No NR SDK calls found.[/dim]")
    for w in result.warnings:
        console.print(f"[yellow]{w}[/yellow]")


cli.add_command(agents_cmd, "agents")
cli.add_command(scan_instrumentation_cmd, "scan-instrumentation")


@click.command("archive")
@click.option("--account", "account_id", required=True, help="NR account id.")
@click.option("--since", "since", required=True, help="NRQL SINCE value, e.g. '7 days ago' or an ISO timestamp.")
@click.option("--until", "until", default=None, help="Optional UNTIL value.")
@click.option("--output", "output_dir", default="./nrdb-archive", type=click.Path())
@click.option(
    "--event-types", default=None,
    help="Comma-separated event types (defaults to the standard NR inventory).",
)
def archive_cmd(account_id: str, since: str, until: Optional[str], output_dir: str, event_types: Optional[str]):
    """Pre-decommission snapshot of NRDB event data (archive-only).

    Produces one JSONL file per event type plus a manifest.json. Resumable
    via `<EventType>.cursor.json` files.
    """
    from tools.nrdb_archive import DEFAULT_EVENT_TYPES, NRDBArchive

    try:
        settings = get_settings()
    except Exception as e:  # noqa: BLE001
        console.print(f"[red]Configuration error: {e}[/red]")
        sys.exit(1)

    nr_client = NewRelicClient(
        api_key=settings.newrelic.api_key,
        account_id=settings.newrelic.account_id,
        region=settings.newrelic.region,
    )

    def run_query(nrql: str, cursor: Optional[str]) -> Dict[str, Any]:
        return nr_client.execute_nrql(nrql, cursor=cursor)

    types = (
        [t.strip() for t in event_types.split(",")]
        if event_types
        else DEFAULT_EVENT_TYPES
    )
    archiver = NRDBArchive(run_query=run_query, account_id=account_id)
    manifest = archiver.archive(
        since=since, until=until, output_dir=output_dir, event_types=types,
    )
    console.print(f"\n[green]Archive complete: {output_dir}[/green]")
    for etype, count in manifest.per_type_counts.items():
        console.print(f"  {etype}: {count} records")
    if manifest.errors:
        console.print(f"\n[yellow]Errors ({len(manifest.errors)}):[/yellow]")
        for etype, err in manifest.errors.items():
            console.print(f"  {etype}: {err}")


cli.add_command(archive_cmd, "archive")


@click.command("audit")
@click.option(
    "--baseline", "baseline_path", required=True, type=click.Path(exists=True),
    help="Path to a transformed_data JSON (e.g. output/transformed/dynatrace_config.json).",
)
@click.option("--output", "output_path", type=click.Path(), default=None,
              help="Optional path to write the JSON drift report.")
def audit_cmd(baseline_path: str, output_path: Optional[str]):
    """Compare a baseline export against the live DT tenant (Phase 20).

    Detects RENAMED / DELETED / MODIFIED / EXTRA drift. Read-only — does
    not write to Dynatrace. Exits 1 when any drift is found.
    """
    from migration.audit import live_snapshot, run_audit

    try:
        settings_obj = get_settings()
    except Exception as e:  # noqa: BLE001
        console.print(f"[red]Configuration error: {e}[/red]"); sys.exit(1)

    dt = DynatraceClient(
        environment_url=settings_obj.dynatrace.environment_url,
        api_token=settings_obj.dynatrace.api_token,
    )
    report = run_audit(Path(baseline_path), lambda: live_snapshot(dt))

    table = Table(title="Migration drift audit")
    table.add_column("Kind", style="cyan")
    table.add_column("Type", style="white")
    table.add_column("Name", style="white")
    table.add_column("Detail", style="dim")
    for d in report.drifts:
        kind_style = {
            "DELETED": "red", "RENAMED": "yellow",
            "MODIFIED": "magenta", "EXTRA": "blue",
        }.get(d.kind, "white")
        table.add_row(
            f"[{kind_style}]{d.kind}[/{kind_style}]",
            d.entity_type, d.name, d.detail or "",
        )
    if report.drifts:
        console.print(table)
    else:
        console.print("[green]No drift detected.[/green]")

    if output_path:
        Path(output_path).write_text(report.to_json())
        console.print(f"[dim]JSON report -> {output_path}[/dim]")

    sys.exit(1 if report.has_drift() else 0)


cli.add_command(audit_cmd, "audit")


if __name__ == "__main__":
    cli()
